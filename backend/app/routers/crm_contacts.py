# NOTICE: This file is protected under RCF-PL
import io
from datetime import datetime, timezone

import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.activity import Activity
from app.models.contact import Contact
from app.models.deal import Deal
from app.models.user import User
from app.schemas.crm import ActivityResponse, ContactCreate, ContactResponse, ContactUpdate, DealResponse
from app.security import get_current_user

router = APIRouter(prefix="/crm/contacts", tags=["crm"])


# [RCF:PROTECTED]
@router.get("", response_model=list[ContactResponse])
# [RCF:PROTECTED]
async def list_contacts(
    search: str | None = None,
    tag: str | None = None,
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    q = select(Contact).where(Contact.user_id == user.id)
    if search:
        q = q.where(Contact.name.ilike(f"%{search}%") | Contact.email.ilike(f"%{search}%") | Contact.company.ilike(f"%{search}%"))
    result = await db.execute(q.order_by(Contact.updated_at.desc()))
    contacts = result.scalars().all()
    if tag:
        contacts = [c for c in contacts if c.tags and tag in c.tags]
    return contacts


# [RCF:PROTECTED]
@router.post("", response_model=ContactResponse, status_code=201)
# [RCF:PROTECTED]
async def create_contact(body: ContactCreate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    contact = Contact(user_id=user.id, **body.model_dump())
    db.add(contact)
    await db.commit()
    await db.refresh(contact)
    return contact


# ── Excel Import ──────────────────────────────────────────────────────────────
# IMPORTANT: /import and /export MUST be defined BEFORE /{contact_id}
# so FastAPI does not try to cast "import"/"export" to int → 422 error.

# [RCF:PROTECTED]
@router.post("/import", status_code=201)
# [RCF:PROTECTED]
async def import_contacts_from_excel(
    file: UploadFile = File(...),
    name_col: str = Query("name"),
    email_col: str = Query("email"),
    phone_col: str = Query("phone"),
    company_col: str = Query("company"),
    tags_col: str = Query("tags"),
    notes_col: str = Query("notes"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    POST /api/crm/contacts/import
    Upload .xlsx / .xls → parse → create contacts with column mapping.

    Query params let the frontend pass custom column names from the user's file:
      ?name_col=Full+Name&email_col=Email+Address&company_col=Organization
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are supported")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse Excel file: {e}")

    df.columns = df.columns.str.strip()

    def _get(row, col: str):
        if col in row.index and pd.notna(row[col]):
            return str(row[col]).strip() or None
        return None

    created, skipped = 0, 0
    for _, row in df.iterrows():
        name = _get(row, name_col)
        if not name:
            skipped += 1
            continue

        email = _get(row, email_col)

        if email:
            existing = await db.execute(
                select(Contact).where(Contact.user_id == user.id, Contact.email == email)
            )
            if existing.scalar_one_or_none():
                skipped += 1
                continue

        raw_tags = _get(row, tags_col)
        tags = [t.strip() for t in raw_tags.split(",") if t.strip()] if raw_tags else None

        contact = Contact(
            user_id=user.id,
            name=name,
            email=email,
            phone=_get(row, phone_col),
            company=_get(row, company_col),
            tags=tags,
            notes=_get(row, notes_col),
            source="excel_import",
        )
        db.add(contact)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "total_rows": len(df)}


# ── Excel Export ──────────────────────────────────────────────────────────────

# [RCF:PROTECTED]
@router.get("/export")
# [RCF:PROTECTED]
async def export_contacts_to_excel(
    tag: str | None = Query(None),
    source: str | None = Query(None),
    created_after: datetime | None = Query(None),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    GET /api/crm/contacts/export
    Returns a styled .xlsx file with all (filtered) contacts.

    Optional filters: ?tag=vip&source=excel_import&created_after=2026-01-01
    """
    q = select(Contact).where(Contact.user_id == user.id)
    if source:
        q = q.where(Contact.source == source)
    if created_after:
        q = q.where(Contact.created_at >= created_after)
    result = await db.execute(q.order_by(Contact.created_at.desc()))
    contacts = result.scalars().all()

    if tag:
        contacts = [c for c in contacts if c.tags and tag in c.tags]

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F2937")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Name", "Email", "Phone", "Company", "Tags", "Source", "Notes", "Created At"]
    col_widths = [6, 25, 30, 18, 25, 20, 15, 35, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    fill_even = PatternFill("solid", start_color="F9FAFB")
    for row_idx, c in enumerate(contacts, start=2):
        fill = fill_even if row_idx % 2 == 0 else None
        values = [
            c.id,
            c.name,
            c.email or "",
            c.phone or "",
            c.company or "",
            ", ".join(c.tags) if c.tags else "",
            c.source or "",
            c.notes or "",
            c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A2"

    summary_row = len(contacts) + 3
    ws.cell(row=summary_row, column=1, value=f"Total: {len(contacts)} contacts").font = Font(italic=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"contacts_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── CRUD by ID (must come AFTER /import and /export) ─────────────────────────

# [RCF:PROTECTED]
@router.get("/{contact_id}", response_model=ContactResponse)
# [RCF:PROTECTED]
async def get_contact(contact_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id, Contact.user_id == user.id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    return contact


# [RCF:PROTECTED]
@router.put("/{contact_id}", response_model=ContactResponse)
# [RCF:PROTECTED]
async def update_contact(contact_id: int, body: ContactUpdate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id, Contact.user_id == user.id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    for key, value in body.model_dump(exclude_unset=True).items():
        setattr(contact, key, value)
    await db.commit()
    await db.refresh(contact)
    return contact


# [RCF:PROTECTED]
@router.delete("/{contact_id}", status_code=204)
# [RCF:PROTECTED]
async def delete_contact(contact_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Contact).where(Contact.id == contact_id, Contact.user_id == user.id))
    contact = result.scalar_one_or_none()
    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")
    await db.delete(contact)
    await db.commit()


# [RCF:PROTECTED]
@router.get("/{contact_id}/activities", response_model=list[ActivityResponse])
# [RCF:PROTECTED]
async def contact_activities(contact_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Activity)
        .where(Activity.contact_id == contact_id, Activity.user_id == user.id)
        .order_by(Activity.created_at.desc())
    )
    return result.scalars().all()


# [RCF:PROTECTED]
@router.get("/{contact_id}/deals", response_model=list[DealResponse])
# [RCF:PROTECTED]
async def contact_deals(contact_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Deal)
        .where(Deal.contact_id == contact_id, Deal.user_id == user.id)
        .order_by(Deal.updated_at.desc())
    )
    return result.scalars().all()
