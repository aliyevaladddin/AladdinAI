# NOTICE: This file is protected under RCF-PL
"""
Excel Reports Generator
GET /api/reports/excel?type=all|deals|contacts|activities
Returns a styled multi-sheet .xlsx report for the authenticated user.
"""
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.activity import Activity
from app.models.contact import Contact
from app.models.deal import Deal
from app.models.user import User
from app.security import get_current_user

router = APIRouter(prefix="/reports", tags=["reports"])

# ── Shared styles ─────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="111827")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", start_color="374151")
SUBHEADER_FONT = Font(bold=True, color="F9FAFB", size=10)
EVEN_FILL = PatternFill("solid", start_color="F3F4F6")
ACCENT_FONT = Font(bold=True, color="1D4ED8")


def _write_header(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = title
    cell.font = Font(bold=True, size=14, color="111827")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"].value = subtitle
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")


def _write_col_headers(ws, row: int, columns: list[tuple[str, int]]) -> None:
    for col_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[row].height = 20


# ── Deals sheet ───────────────────────────────────────────────────────────────

async def _build_deals_sheet(ws, user_id: int, db: AsyncSession) -> int:
    _write_header(
        ws,
        "AladdinAI — Deals Report",
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
    )
    _write_col_headers(ws, 4, [
        ("ID", 6), ("Title", 30), ("Contact", 25), ("Stage", 15),
        ("Amount ($)", 14), ("Currency", 10), ("Probability", 12), ("Created At", 18),
    ])

    result = await db.execute(
        select(Deal).where(Deal.user_id == user_id).order_by(Deal.created_at.desc())
    )
    deals = result.scalars().all()

    # Batch-load contacts for name lookup
    contact_ids = {d.contact_id for d in deals if d.contact_id}
    contacts_map: dict[int, str] = {}
    if contact_ids:
        cr = await db.execute(select(Contact).where(Contact.id.in_(contact_ids)))
        contacts_map = {c.id: c.name for c in cr.scalars().all()}

    total_amount = 0.0
    for row_idx, deal in enumerate(deals, start=5):
        fill = EVEN_FILL if row_idx % 2 == 0 else None
        amount = deal.amount or 0.0
        total_amount += amount
        values = [
            deal.id,
            deal.title,
            contacts_map.get(deal.contact_id, "—"),
            deal.stage or "—",
            amount,
            deal.currency or "USD",
            f"{deal.probability}%",
            deal.created_at.strftime("%Y-%m-%d"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            if col_idx == 5:
                cell.number_format = "#,##0.00"

    # Summary row
    summary_row = len(deals) + 6
    ws.cell(row=summary_row, column=4, value="TOTAL").font = Font(bold=True)
    total_cell = ws.cell(row=summary_row, column=5, value=total_amount)
    total_cell.font = ACCENT_FONT
    total_cell.number_format = "#,##0.00"
    ws.freeze_panes = "A5"
    return len(deals)


# ── Contacts sheet ────────────────────────────────────────────────────────────

async def _build_contacts_sheet(ws, user_id: int, db: AsyncSession) -> int:
    _write_header(
        ws,
        "AladdinAI — Contacts Report",
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
    )
    _write_col_headers(ws, 4, [
        ("ID", 6), ("Name", 28), ("Email", 30), ("Phone", 18),
        ("Company", 25), ("Tags", 20), ("Source", 15), ("Created At", 18),
    ])

    result = await db.execute(
        select(Contact).where(Contact.user_id == user_id).order_by(Contact.created_at.desc())
    )
    contacts = result.scalars().all()

    for row_idx, c in enumerate(contacts, start=5):
        fill = EVEN_FILL if row_idx % 2 == 0 else None
        values = [
            c.id,
            c.name,
            c.email or "—",
            c.phone or "—",
            c.company or "—",
            ", ".join(c.tags) if c.tags else "—",
            c.source or "—",
            c.created_at.strftime("%Y-%m-%d"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A5"
    return len(contacts)


# ── Activities sheet ──────────────────────────────────────────────────────────

async def _build_activities_sheet(ws, user_id: int, db: AsyncSession) -> int:
    _write_header(
        ws,
        "AladdinAI — Activities Report",
        f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC",
    )
    _write_col_headers(ws, 4, [
        ("ID", 6), ("Type", 15), ("Channel", 15), ("Subject", 35),
        ("Contact", 25), ("Deal ID", 10), ("Created At", 18),
    ])

    result = await db.execute(
        select(Activity).where(Activity.user_id == user_id).order_by(Activity.created_at.desc())
    )
    activities = result.scalars().all()

    contact_ids = {a.contact_id for a in activities if a.contact_id}
    contacts_map: dict[int, str] = {}
    if contact_ids:
        cr = await db.execute(select(Contact).where(Contact.id.in_(contact_ids)))
        contacts_map = {c.id: c.name for c in cr.scalars().all()}

    for row_idx, a in enumerate(activities, start=5):
        fill = EVEN_FILL if row_idx % 2 == 0 else None
        values = [
            a.id,
            a.type or "—",
            a.channel or "—",
            (a.subject or "—")[:80],
            contacts_map.get(a.contact_id, "—"),
            a.deal_id or "—",
            a.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    ws.freeze_panes = "A5"
    return len(activities)


# ── Endpoint ──────────────────────────────────────────────────────────────────

# [RCF:PROTECTED]
@router.get("/excel")
# [RCF:PROTECTED]
async def download_excel_report(
    type: str = Query("all", enum=["all", "deals", "contacts", "activities"]),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    GET /api/reports/excel?type=all|deals|contacts|activities
    Returns a styled multi-sheet Excel workbook for the authenticated user.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    if type in ("all", "deals"):
        await _build_deals_sheet(wb.create_sheet("Deals"), user.id, db)

    if type in ("all", "contacts"):
        await _build_contacts_sheet(wb.create_sheet("Contacts"), user.id, db)

    if type in ("all", "activities"):
        await _build_activities_sheet(wb.create_sheet("Activities"), user.id, db)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"aladdinai_report_{type}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
